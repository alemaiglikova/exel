import openpyxl

wb = openpyxl.load_workbook("Excel.xlsx")
ws = wb.active
massive= []
max_row = ws.max_row
max_column = ws.max_column

for row in range(1, max_row + 1):
    massive2 = []
    for column in range(1, max_column + 1):
        massive2.append(ws.cell(row, column).value)
    massive.append(massive2)